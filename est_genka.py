# -*- coding: utf-8 -*-
"""
見積書xlsxの読み込み(ruiji.py相当)から、Excel VBAマクロ「仕分け」相当の
仕分・集計処理までを一気通貫で行い、仕分済みのxlsxを出力する。
"""
import os
import sys

import pandas as pd
import pyodbc
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import column_index_from_string

from dnd_filepicker import pick_file

# ---------------------------------------------------------------------------
# 設定
# ---------------------------------------------------------------------------
SHEET = "内訳書"

DB_PATH = r"\\192.168.1.136\wwwroot\見積物件管理1.accdb"

TITLE_ROWS = 2
DATA_ROWS = 22
FOOTER_ROWS = 1
PAGE_SIZE = TITLE_ROWS + DATA_ROWS + FOOTER_ROWS

DATA_START_ROW = 3  # VBAが処理対象とする最初の行(1行目=見出し, 2行目=未使用)

SAME_AS_ABOVE_ITEM = "同上継手接合材　　　　　　　　　"
GENKA_RATE = 0.15742
MARGIN_TARGET = 0.2

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# 見積書の読み込み・ページ解析 (ruiji.py と同じロジック)
# ---------------------------------------------------------------------------
def parse_quote(input_file, sheet):
    df = pd.read_excel(input_file, sheet_name=sheet, header=None)

    total_rows = len(df)
    num_pages = total_rows // PAGE_SIZE

    all_data = []
    for i in range(num_pages):
        start = i * PAGE_SIZE + TITLE_ROWS
        end = start + DATA_ROWS
        page_df = df.iloc[start:end]
        page_df = page_df.iloc[:, [0, 7, 8, 9, 10]]
        all_data.append(page_df.reset_index(drop=True))

    result = pd.concat(all_data, ignore_index=True)
    result = result.dropna(how="all")

    result.insert(0, "flag", 0)
    result.insert(3, "flag2", "")
    result.insert(4, "", "")

    condition = result.iloc[:, 7].notna() & result.iloc[:, 2].notna()
    result.loc[condition, "flag"] = 8

    return result


def to_cell_value(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return v


# ---------------------------------------------------------------------------
# シート組み立て
# ---------------------------------------------------------------------------
def make_sheet_title(base_name):
    invalid = set('[]:*?/\\')
    title = "".join(c for c in base_name if c not in invalid).strip()
    return (title[:31] if title else "Sheet1")


def write_data_columns(ws, result):
    """A:I 列にデータを書き込み、行1=見出し、行2以降=データとする。
    戻り値: last (データ最終行番号)"""
    headers = {"A": "flag", "C": 0, "D": 7, "E": "flag2", "G": 8, "H": 9, "I": 10}
    for col, val in headers.items():
        ws[f"{col}1"] = val

    row = 2
    for _, rec in result.iterrows():
        ws.cell(row=row, column=1, value=to_cell_value(rec["flag"]))
        # column=2 (B) は仕分カテゴリ。この時点では空欄のまま。
        ws.cell(row=row, column=3, value=to_cell_value(rec[0]))
        ws.cell(row=row, column=4, value=to_cell_value(rec[7]))
        ws.cell(row=row, column=5, value=to_cell_value(rec["flag2"]))
        # column=6 (F) は常に空列。
        ws.cell(row=row, column=7, value=to_cell_value(rec[8]))
        ws.cell(row=row, column=8, value=to_cell_value(rec[9]))
        ws.cell(row=row, column=9, value=to_cell_value(rec[10]))
        row += 1

    last = row - 1
    return last


def approx_autofit(ws, columns, min_width=6, padding=2):
    """openpyxlにAutoFitは無いため、文字数から幅を概算する。"""
    for col in columns:
        max_len = 0
        for cell in ws[col]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col].width = max(min_width, max_len + padding)


def apply_base_layout(ws):
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["Q"].width = 16
    ws.column_dimensions["R"].width = 10
    ws.column_dimensions["S"].width = 7
    ws.column_dimensions["T"].width = 10
    ws.column_dimensions["U"].width = 7
    ws.column_dimensions["V"].width = 10


def apply_comma_format(ws, column, first_row, last_row):
    for r in range(first_row, last_row + 1):
        ws.cell(row=r, column=column_index_from_string(column)).number_format = "#,##0"


# ---------------------------------------------------------------------------
# 仕分項目入力 (仕分ワード照合 + 「同上継手接合材」の上行コピー)
# ---------------------------------------------------------------------------
def classify_rows(ws, cur, last):
    for m in range(DATA_START_ROW, last + 1):
        flag_val = ws.cell(row=m, column=1).value
        item_val = ws.cell(row=m, column=3).value
        if flag_val == 8 and item_val not in (None, ""):
            word = str(item_val)
            cur.execute(
                "select bunrui from 仕分ワード where word = ?", (word + " ",)
            )
            rec = cur.fetchone()
            if rec is not None:
                ws.cell(row=m, column=2).value = rec.bunrui

    for x in range(DATA_START_ROW, last + 1):
        item_val = ws.cell(row=x, column=3).value
        if item_val == SAME_AS_ABOVE_ITEM:
            ws.cell(row=x, column=2).value = ws.cell(row=x - 1, column=2).value


def add_estimate_total(ws, last):
    range1 = f"$A${DATA_START_ROW}:$A${last}"
    range3 = f"$I${DATA_START_ROW}:$I${last}"
    ws.cell(row=last + 1, column=9, value=f"=SUMIF({range1},8,{range3})")
    apply_comma_format(ws, "I", last + 1, last + 1)


# ---------------------------------------------------------------------------
# 仕訳表作成 (K:P)
# ---------------------------------------------------------------------------
def build_bunrui_table(ws, cur, last):
    ws["M2"] = "見積"
    ws["N2"] = "仕切"
    ws["O2"] = "掛率"

    cur.execute(
        "select nb, bunrui, fukuri from 仕分A "
        "union all select nb, bunrui, fukuri from 仕分B "
        "union all select nb, bunrui, fukuri from 仕分C "
        "union all select nb, bunrui, fukuri from 仕分D "
        "order by nb asc"
    )
    records = cur.fetchall()

    # VBAはCells(n,13)の数式評価結果(SUMIF結果)を見て分岐している。
    # openpyxlは数式を評価しないため、同じ判定をPython側の集計で代替する。
    sum_by_bunrui = {}
    for r in range(DATA_START_ROW, last + 1):
        cat = ws.cell(row=r, column=2).value
        amt = ws.cell(row=r, column=9).value
        if cat:
            sum_by_bunrui[cat] = sum_by_bunrui.get(cat, 0) + (amt or 0)

    range2 = f"$B${DATA_START_ROW}:$B${last}"
    range3 = f"$I${DATA_START_ROW}:$I${last}"

    n = 3
    for rec in records:
        nb, bunrui, fukuri = rec.nb, rec.bunrui, rec.fukuri
        ws.cell(row=n, column=11, value=nb)
        ws.cell(row=n, column=12, value=bunrui)
        l_addr = f"$L${n}"
        ws.cell(row=n, column=13, value=f"=SUMIF({range2},{l_addr},{range3})")

        m_value = sum_by_bunrui.get(bunrui, 0)
        if m_value == 0:
            ws.cell(row=n, column=15, value="")
        else:
            ws.cell(row=n, column=15, value=f"=$N${n}/$M${n}")

        nx = fukuri if fukuri is not None else 0
        if isinstance(nx, float) and nx.is_integer():
            nx = int(nx)
        ws.cell(row=n, column=16, value=f"=$M${n}*{nx}")
        n += 1

    hlast = n - 1

    apply_comma_format(ws, "M", DATA_START_ROW, hlast)
    apply_comma_format(ws, "N", DATA_START_ROW, hlast)

    ws.cell(row=hlast + 2, column=12, value="粗利")
    ws.cell(row=hlast + 3, column=12, value="積算NET")
    ws.cell(row=hlast + 1, column=13, value=f"=SUM($M$3:$M${hlast})")
    ws.cell(row=hlast + 1, column=14, value=f"=SUM($N$3:$N${hlast})")
    ws.cell(row=hlast + 1, column=16, value=f"=SUM($P$3:$P${hlast})*{GENKA_RATE}")
    ws.cell(row=hlast + 2, column=16, value=MARGIN_TARGET)
    ws.cell(row=hlast + 2, column=16).number_format = "0%"
    ws.cell(row=hlast + 2, column=14, value=f"=$N${hlast + 3}-$N${hlast + 1}")
    ws.cell(row=hlast + 3, column=14, value=f"=$N${hlast + 1}/(1-$P${hlast + 2})")

    apply_comma_format(ws, "N", hlast + 1, hlast + 3)
    apply_comma_format(ws, "M", hlast + 1, hlast + 1)

    for row in ws[f"K2:O{hlast + 3}"]:
        for cell in row:
            cell.border = THIN_BORDER

    approx_autofit(ws, ["L"])

    return hlast


def add_bunrui_validation(wb, ws, last, hlast):
    defined_name = DefinedName("仕分", attr_text=f"'{ws.title}'!$L$3:$L${hlast}")
    wb.defined_names["仕分"] = defined_name

    dv = DataValidation(type="list", formula1="=仕分", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{DATA_START_ROW}:B{last}")


# ---------------------------------------------------------------------------
# 機器集計表 (Q:V)
# ---------------------------------------------------------------------------
def build_equipment_summary(ws):
    ws["R3"] = "定価"
    ws["S3"] = "見積"
    ws["U3"] = "仕切"
    ws.merge_cells("S3:T3")
    ws.merge_cells("U3:V3")

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws["S3"].alignment = center
    ws["U3"].alignment = center

    for row in ws["Q3:V15"]:
        for cell in row:
            cell.border = THIN_BORDER

    for mx in range(4, 16):
        ws.cell(row=mx, column=20, value=f"=R{mx}*S{mx}")
        ws.cell(row=mx, column=22, value=f"=R{mx}*U{mx}")

    apply_comma_format(ws, "R", 4, 15)
    apply_comma_format(ws, "T", 4, 15)
    apply_comma_format(ws, "V", 4, 15)


# ---------------------------------------------------------------------------
# メイン処理
# ---------------------------------------------------------------------------
def validate_input_file(path):
    if not path.lower().endswith(".xlsx"):
        return "xlsxファイルを選択してください。"
    return None


def pick_input_file():
    return pick_file(
        filetypes=[("Excel files", "*.xlsx")],
        title="見積書ファイル選択",
        prompt="ここに見積書(.xlsx)を\nドラッグ&ドロップ",
        validate=validate_input_file,
    )



def run(input_file):
    result = parse_quote(input_file, SHEET)

    wb = Workbook()
    ws = wb.active
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    ws.title = make_sheet_title(base_name)

    last = write_data_columns(ws, result)
    apply_base_layout(ws)
    approx_autofit(ws, ["A", "F", "G"])
    apply_comma_format(ws, "H", 2, last)
    apply_comma_format(ws, "I", 2, last)

    add_estimate_total(ws, last)

    conn = pyodbc.connect(
        r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + DB_PATH + ";"
    )
    try:
        cur = conn.cursor()
        classify_rows(ws, cur, last)
        hlast = build_bunrui_table(ws, cur, last)
    finally:
        conn.close()

    add_bunrui_validation(wb, ws, last, hlast)
    build_equipment_summary(ws)

    output_file = os.path.splitext(input_file)[0] + "(仕訳).xlsx"
    wb.save(output_file)
    return output_file


def main():
    if len(sys.argv) > 1:
        # アイコンへのドラッグ&ドロップ起動: 渡されたファイルをそのまま使う
        input_file = sys.argv[1]
        error = validate_input_file(input_file)
        if error:
            messagebox.showerror("エラー", error)
            return
    else:
        input_file = pick_input_file()
        if not input_file:
            return

    try:
        output_file = run(input_file)
    except Exception as e:
        messagebox.showerror("エラー", f"処理に失敗しました。\n{e}")
        raise

    print("saved:", output_file)
    messagebox.showinfo("完了", f"作成しました:\n{output_file}")


if __name__ == "__main__":
    main()
